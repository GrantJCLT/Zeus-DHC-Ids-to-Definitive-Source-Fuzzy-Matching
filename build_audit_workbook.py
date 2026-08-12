#!/usr/bin/env python3
"""Build the branded Accuracy Audit workbook from a scored run.

The format follows Zeus_DHC_ID_Accuracy_Audit.xlsx, which is kept as the
reporting template. Two deliberate additions over that original:

  * a Geo_Conflict sheet, which post-dates the template and is the
    highest-value review population;
  * Unreferenced_Definitive is generated rather than hand-built.

  py build_audit_workbook.py --scored audit_2026_08_all4_scored.csv \
      --config sources.yaml --out Zeus_DHC_ID_Accuracy_Audit_2026_08.xlsx
"""
import argparse
import os

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from dhc_match_v2 import load_config, load_definitive, load_locations, \
    locations_as_identity, _read_roles

# Brand palette, lifted from the template.
NAVY, PURPLE = 'FF004065', 'FF8247AF'
INK, GREY = 'FF2D2A26', 'FF959492'
CREAM, LILAC = 'FFFEF9E5', 'FFEFE7F7'

H1 = Font(size=16, bold=True, color='FFFFFFFF')
H2 = Font(size=11, color='FFFFFFFF')
SECTION = Font(size=13, bold=True, color=INK)
HEAD = Font(size=10, bold=True, color='FFFFFFFF')
BODY = Font(size=10, color=INK)
NOTE = Font(size=8, color=GREY)
FILL_NAVY = PatternFill('solid', start_color=NAVY)
FILL_PURPLE = PatternFill('solid', start_color=PURPLE)
FILL_CREAM = PatternFill('solid', start_color=CREAM)
FILL_LILAC = PatternFill('solid', start_color=LILAC)

WRAP = Alignment(wrap_text=True, vertical='top')


def _put(ws, ref, value, font=BODY, fill=None, fmt=None, align=None):
    c = ws[ref]
    c.value = value
    c.font = font
    if fill:
        c.fill = fill
    if fmt:
        c.number_format = fmt
    if align:
        c.alignment = align
    return c


def _table(ws, top_row, headers, rows, widths=None, notes_col=None):
    """Header band + body rows, returns the next free row."""
    for j, h in enumerate(headers):
        _put(ws, f'{get_column_letter(2 + j)}{top_row}', h, HEAD, FILL_PURPLE)
    r = top_row + 1
    for row in rows:
        for j, v in enumerate(row):
            col = get_column_letter(2 + j)
            font = NOTE if (notes_col is not None and j == notes_col) else BODY
            fmt = None
            if isinstance(v, float) and 0 <= v <= 1 and j and notes_col != j:
                fmt = '0.0%'
            elif isinstance(v, (int, np.integer)):
                fmt = '#,##0'
            _put(ws, f'{col}{r}', v, font, None, fmt,
                 WRAP if font is NOTE else None)
        r += 1
    if widths:
        for j, w in enumerate(widths):
            ws.column_dimensions[get_column_letter(2 + j)].width = w
    return r + 1


def sheet_data(wb, title, df, widths=None):
    """A tabular sheet: purple header, frozen top row, autofilter."""
    ws = wb.create_sheet(title[:31])
    ws.append(list(df.columns))
    for c in ws[1]:
        c.font, c.fill = HEAD, FILL_PURPLE
    for row in df.itertuples(index=False):
        ws.append(['' if (isinstance(v, float) and np.isnan(v)) else v
                   for v in row])
    ws.freeze_panes = 'A2'
    if len(df):
        ws.auto_filter.ref = f'A1:{get_column_letter(len(df.columns))}{len(df) + 1}'
    for j, col in enumerate(df.columns, start=1):
        w = 46 if 'Name' in str(col) else (
            26 if 'Address' in str(col) or str(col).endswith('Id') else 15)
        ws.column_dimensions[get_column_letter(j)].width = w
    return ws


def build_summary(wb, s, funnel, verdicts, by_type, by_pop, gains, subtitle):
    ws = wb.create_sheet('Summary')
    for col, w in zip('ABCDE', (4, 62, 16, 14, 60)):
        ws.column_dimensions[col].width = w

    ws.merge_cells('A1:E1')
    _put(ws, 'A1', 'Jackson and Coker Locum Tenens - Zeus DHC Identifier '
                   'Accuracy Audit', H1, FILL_NAVY)
    ws.merge_cells('A2:E2')
    _put(ws, 'A2', subtitle, H2, FILL_PURPLE)

    _put(ws, 'B4', 'Headline answer', SECTION)
    ws.merge_cells('B5:E7')
    total = funnel['testable']
    corr = verdicts.get('ID corroborated', 0)
    _put(ws, 'B5',
         f'Of the {total:,} Zeus entities whose DHC identifier can actually be '
         f'tested against a Definitive record, {corr:,} ({corr / total:.1%}) '
         f'are corroborated by name and address, and '
         f'{funnel["corr_or_prob"]:,} ({funnel["corr_or_prob"] / total:.1%}) '
         f'are corroborated or probable. {verdicts.get("Likely wrong ID", 0):,} '
         f'look like the wrong identifier. This is a measure of the '
         f'{total:,} testable entities, not of all {funnel["zeus"]:,} Zeus '
         f'entities carrying an identifier - see the population funnel below.',
         BODY, FILL_CREAM, align=WRAP)

    r = 9
    _put(ws, f'B{r}', '1. Population funnel', SECTION)
    r = _table(ws, r + 1, ['Measure', 'Rows', '% of Zeus', 'Reading'],
               funnel['rows'], notes_col=3)
    # shade the supplied-input row, as the template does
    for col in 'BCDE':
        ws[f'{col}11'].fill = FILL_LILAC
    _put(ws, f'B{r}', 'Shaded row is a direct input from the query. All other '
                      'counts are derived.', NOTE)

    r += 2
    _put(ws, f'B{r}', '2. Verdict on the testable population', SECTION)
    _put(ws, f'B{r + 1}', f'Denominator is {total:,} rows whose identifier '
                          f'resolves to a Definitive record.', NOTE)
    r = _table(ws, r + 2, ['Verdict', 'Rows', 'Share', 'Reading'],
               [[k, int(v), v / total, note] for k, v, note in verdicts['rows']],
               notes_col=3)

    _put(ws, f'B{r}', '3. By Definitive entity type', SECTION)
    r = _table(ws, r + 1, ['Entity type', 'Rows', 'Corroborated', 'Share'],
               by_type)

    if by_pop:
        _put(ws, f'B{r}', '4. By Zeus population', SECTION)
        _put(ws, f'B{r + 1}', 'Populations overlap - an entity that is both a '
                              'client and a work location is counted in each, '
                              'so these rows sum to more than the testable '
                              'total. Its names and addresses from every '
                              'population are pooled into one verdict.', NOTE)
        r = _table(ws, r + 2,
                   ['Zeus population', 'Testable', 'Corroborated', 'Share'],
                   by_pop)

    _put(ws, f'B{r}', '5. What the enrichment contributed', SECTION)
    _put(ws, f'B{r + 1}', 'Two changes since the baseline run, so this delta is '
                          'their combined effect: Definitive service locations '
                          'mean a Zeus address is compared against every known '
                          'site rather than the HQ alone, and pooling means an '
                          'entity is compared using its name and address from '
                          'every Zeus population it belongs to.', NOTE)
    r = _table(ws, r + 2, ['Measure', 'HQ only', 'With locations', 'Reading'],
               gains, notes_col=3)
    return ws


def build_methodology(wb, notes):
    ws = wb.create_sheet('Methodology')
    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 118
    _put(ws, 'A1', 'Methodology and assumptions', SECTION)
    r = 2
    for head, text in notes:
        _put(ws, f'B{r}', head, Font(size=10, bold=True, color=INK),
             align=WRAP)
        _put(ws, f'C{r}', text, BODY, align=WRAP)
        r += 1
    return ws


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--scored', required=True)
    ap.add_argument('--config', default='sources.yaml')
    ap.add_argument('--baseline', help='earlier HQ-only scored csv, for the '
                                      'location-gain comparison')
    ap.add_argument('--out', required=True)
    a = ap.parse_args()

    s = pd.read_csv(a.scored, low_memory=False)
    prefix = a.scored.replace('_scored.csv', '')
    extract = f'{prefix}_zeus_extract.csv'
    unver = f'{prefix}_unverifiable.csv'
    z = pd.read_csv(extract, low_memory=False) if os.path.exists(extract) else None
    u = pd.read_csv(unver, low_memory=False) if os.path.exists(unver) else None

    cfg = load_config(a.config)
    # The extract holds one row per population per entity; the audit's grain is
    # the entity, so the funnel must count distinct EntityIds, not extract rows.
    zeus_rows = (int(z.EntityId.nunique()) if z is not None and 'EntityId' in z
                 else len(s) + (len(u) if u is not None else 0))
    pop_rows = len(z) if z is not None else None
    testable = len(s)
    ident_ids = set()
    for b in cfg['definitive']:
        ident_ids |= set(pd.to_numeric(_read_roles(b, ('id',))[b['id']],
                                       errors='coerce').dropna().astype('int64'))
    maxdig = max(len(str(i)) for i in ident_ids)
    allids = pd.to_numeric(s.DHC_Id, errors='coerce')
    if u is not None and 'DHC_Id' in u.columns:
        allids = pd.concat([allids, pd.to_numeric(u.DHC_Id, errors='coerce')])

    # Count a row implausible if ANY of its identifier columns is too wide -
    # not just the resolved one. An NPI-shaped LEVS value on a row whose Entity
    # id is populated never becomes the resolved id, but is still a defect.
    def _too_wide(series):
        v = pd.to_numeric(series, errors='coerce').dropna().astype('int64')
        return v.astype(str).str.len() > maxdig

    implausible = 0
    if z is not None:
        flags = None
        for c in (cfg['zeus'].get('id_columns') or []):
            if c in z.columns:
                f = _too_wide(z[c]).reindex(z.index, fill_value=False)
                flags = f if flags is None else (flags | f)
        if flags is not None:
            # One row per population per entity - count each entity once.
            implausible = int(z.assign(_f=flags).groupby('EntityId')._f.any().sum())

    vc = s.Verdict.value_counts().to_dict()
    corr_or_prob = int(s.Verdict.str.startswith(('ID corroborated',
                                                 'Probable')).sum())
    funnel = {
        'zeus': zeus_rows, 'testable': testable, 'corr_or_prob': corr_or_prob,
        'rows': [
            ['Zeus entities supplied', zeus_rows, 1.0,
             f'Row grain: one per EntityId. Pooled from '
             f'{pop_rows:,} population rows across six queries; an entity in '
             f'several populations is counted once.' if pop_rows else
             'Row grain: one per EntityId.'],
            ['DHC identifier populated', int(allids.notna().sum()),
             float(allids.notna().sum()) / zeus_rows,
             'Entity_DHC_VerifiedSourceId, falling back to '
             'LEVS_DHC_VerifiedSourceId.'],
            ['Identifier blank', int(zeus_rows - allids.notna().sum()),
             float(zeus_rows - allids.notna().sum()) / zeus_rows,
             'The query only returns rows carrying an identifier.'],
            ['Identifier format implausible', implausible,
             implausible / zeus_rows,
             f'Rows where either identifier column exceeds {maxdig} digits, '
             f'the widest id in the reference set - typically NPI-shaped '
             f'values in LEVS_DHC_VerifiedSourceId. Counted across both '
             f'columns, so it includes rows whose resolved id is fine.'],
            ['Identifier resolves to a Definitive record', testable,
             testable / zeus_rows,
             'The testable population. Everything below is scored on these.'],
            ['Identifier not in the reference set',
             zeus_rows - testable, (zeus_rows - testable) / zeus_rows,
             f'Unverifiable, not wrong. Full list in '
             f'{os.path.basename(prefix)}_unverifiable.csv.'],
        ]}

    reading = {
        'ID corroborated': 'Name and address agree. No action.',
        'Probable - name agrees, address differs':
            'Name matches; no known location matches the Zeus address.',
        'Probable - address agrees, name differs':
            'Address matches a known site; the name does not.',
        'Needs review': 'Neither side is convincing. Human judgement needed.',
        'Likely wrong ID': 'Strong signal the identifier points elsewhere.',
    }
    verdicts = dict(vc)
    verdicts['rows'] = [[k, int(v), reading.get(k, '')]
                        for k, v in sorted(vc.items(), key=lambda x: -x[1])]

    g = s.groupby('DHC_Entity_Type').agg(
        Rows=('Verdict', 'size'),
        Corroborated=('Verdict', lambda x: int((x == 'ID corroborated').sum())))
    by_type = [[i, int(r.Rows), int(r.Corroborated), r.Corroborated / r.Rows]
               for i, r in g.iterrows()]

    by_pop = []
    if 'Zeus_Sources' in s.columns:
        labels = sorted({x for v in s.Zeus_Sources.dropna()
                         for x in str(v).split('|')})
        for lab in labels:
            m = s.Zeus_Sources.fillna('').str.split('|').map(
                lambda v, l=lab: l in v)
            n_ = int(m.sum())
            ok = int((s[m].Verdict == 'ID corroborated').sum())
            by_pop.append([lab, n_, ok, ok / n_ if n_ else 0.0])

    gains = []
    if a.baseline and os.path.exists(a.baseline):
        b = pd.read_csv(a.baseline, low_memory=False)
        # Restricted to entities present in BOTH runs. The baseline predates the
        # extra Zeus populations, so comparing full totals would credit the
        # location data with a population increase it had nothing to do with.
        keep = set(b.EntityId) & set(s.EntityId)
        bb, ss = b[b.EntityId.isin(keep)], s[s.EntityId.isin(keep)]
        note = f'Like-for-like on the {len(keep):,} entities in both runs.'

        def pair(label, bser, sser, reading):
            return [label, int(bser.sum()), int(sser.sum()), reading]

        gains = [
            pair('Corroborated', bb.Verdict == 'ID corroborated',
                 ss.Verdict == 'ID corroborated', note),
            pair('Address divergent (corroborated, addr<60)',
                 (bb.Verdict == 'ID corroborated') & (bb.Address_Score < 60),
                 (ss.Verdict == 'ID corroborated') & (ss.Address_Score < 60),
                 'Zeus address matches neither HQ nor any known site.'),
            pair('Geo conflict',
                 bb.Geo_Conflict if 'Geo_Conflict' in bb else pd.Series(dtype=bool),
                 ss.Geo_Conflict if 'Geo_Conflict' in ss else pd.Series(dtype=bool),
                 'Now means no known location is in the Zeus state.'),
            pair('Address matched a satellite, not HQ',
                 pd.Series([False] * len(bb)),
                 ss.Address_Match_Source == 'Location'
                 if 'Address_Match_Source' in ss else pd.Series([False] * len(ss)),
                 'Impossible before service locations were available.'),
        ]

    subtitle = ('Six Zeus populations (Client, Work Location, Health System, '
                'GPO, Agency, VMS) vs. Definitive Hospital, Physician Group, '
                'GPO and Practice Location exports  |  read live from the '
                'read-only replica')

    wb = Workbook()
    wb.remove(wb.active)
    build_summary(wb, s, funnel, verdicts, by_type, by_pop, gains, subtitle)
    build_methodology(wb, [
        ('Source of truth', 'Zeus is read live from the failover replica with '
         'ApplicationIntent=ReadOnly; the run asserts the database is '
         'READ_ONLY. Every run snapshots its exact input to '
         '<out>_zeus_extract.csv.'),
        ('Six populations, pooled', 'Zeus is read through six queries - '
         'IsClient, IsWorkLocation, IsHealthSystem, IsGPO, IsAgency, IsVMS - '
         'each reading its own *Info table and so its own column names. 6,855 '
         'of 12,803 entities appear in more than one, holding a different name '
         'and address in each. Neither is authoritative, so all of them become '
         'candidates and the best match wins: one verdict per entity, no double '
         'counting. Zeus_Sources records which populations contributed.'),
        ('Population', 'The queries return non-archived entities '
         'carrying a DHC identifier, and EXCLUDES entities created by a '
         'Definitive import. That exclusion removes 95.6% of identifier-'
         'carrying rows, whose identifiers are near self-referential. Figures '
         'here are therefore not comparable to any run before 2026-07-31.'),
        ('Identifier resolution', 'Entity_DHC_VerifiedSourceId first, falling '
         'back to LEVS_DHC_VerifiedSourceId. Where both are populated and '
         'disagree the row is flagged on ID_Conflicts.'),
        ('Why the denominator is what it is', 'An identifier absent from the '
         'reference set is unverifiable, not wrong. Never quote an accuracy '
         'rate against all Zeus rows; always state the testable denominator.'),
        ('Reference set', 'Hospital, Physician Group and GPO exports supply '
         'identity - what an id IS. They share no ids.'),
        ('Service locations', 'Practice Locations is a child table keyed on the '
         'PARENT id. It is never an identity source; instead every location '
         'contributes a candidate address, city, state, zip and name alias, so '
         'a Zeus service address is compared against every known site.'),
        ('Normalisation', 'Both sides lowercased, ampersands expanded, '
         'punctuation stripped. Zeus stores full state names, Definitive '
         'stores codes; states are normalised before comparison.'),
        ('Former names matter', 'Definitive embeds former and alternate names '
         'in parentheses ((FKA ...), (AKA ...)). These are parsed out and '
         'scored as aliases, otherwise renamed facilities read as mismatches.'),
        ('Zeus name fields are aliases', 'ClientEntityName and ClientInfoName '
         'are alternates, not primary-and-fallback. Both are scored and the '
         'best match wins.'),
        ('Name outranks address', 'Verdict is driven by name; address is '
         'reported separately. A single blended score would conflate "is the '
         'id right" with "is our address current". Act on Verdict; '
         'Confidence_Score is for sorting and trending only.'),
        ('Name scoring', 'A 35/65 blend of token_set_ratio and '
         'token_sort_ratio. Bare token_set_ratio returns 100 whenever one '
         'token set is a subset of the other, which let short generic names '
         'win spuriously.'),
        ('Geo conflict', 'A strong name match where no known location of the '
         'entity is in the Zeus state. Flagged rather than reclassified: some '
         'are legitimate, some are the wrong id, and string comparison cannot '
         'separate them.'),
        ('Known limit', 'Parent-versus-child entities and same-named entities '
         'in different places. Name agreement proves the two sides mean the '
         'same NAME, not necessarily the same ENTITY. Closing that needs a '
         'human or a third identifier such as NPI.'),
    ])

    show = [c for c in ['EntityId', 'Zeus_Sources', 'Zeus_Name',
                        'Zeus_Names_All', 'Zeus_Address', 'Zeus_City',
                        'Zeus_State', 'Zeus_Zip', 'DHC_Id', 'DHC_Matched_Name',
                        'DHC_Entity_Type', 'DHC_City', 'DHC_State',
                        'Name_Score', 'Address_Score', 'State_Score',
                        'Address_Match_Source', 'Location_Count', 'Verdict',
                        'Suggested_DHC_Id', 'Suggested_Name',
                        'Suggested_Name_Score', 'Correction_Recommended']
            if c in s.columns]

    rq = s[s.Verdict.isin(['Needs review', 'Likely wrong ID'])]
    sheet_data(wb, 'Review_Queue', rq[show].sort_values('Name_Score'))
    if 'Geo_Conflict' in s.columns:
        sheet_data(wb, 'Geo_Conflict',
                   s[s.Geo_Conflict][show].sort_values('Name_Score',
                                                       ascending=False))
    if 'Address_Divergent' in s.columns:
        sheet_data(wb, 'Address_Divergence',
                   s[s.Address_Divergent][show].sort_values('Address_Score'))
    if 'Correction_Recommended' in s.columns:
        cr = s[s.Correction_Recommended.fillna(False).astype(bool)]
        sheet_data(wb, 'Corrections_Recommended', cr[show])

    dup = s[s.DHC_Id.duplicated(keep=False)].sort_values(['DHC_Id', 'EntityId'])
    sheet_data(wb, 'Duplicate_IDs', dup[show])

    if 'DHC_Id_Conflict' in s.columns:
        cf = s[s.DHC_Id_Conflict.fillna(False).astype(bool)]
        cols = [c for c in ['EntityId', 'Zeus_Sources', 'Zeus_Name',
                            'Zeus_City', 'Zeus_State',
                            'Entity_DHC_VerifiedSourceId',
                            'LEVS_DHC_VerifiedSourceId', 'DHC_Id',
                            'DHC_Matched_Name', 'Verdict'] if c in s.columns]
        sheet_data(wb, 'ID_Conflicts', cf[cols])

    sheet_data(wb, 'Scored_Detail', s)

    # Reference records nothing in Zeus points at - generated, not hand-built.
    L = load_locations(cfg.get('locations'))
    d = load_definitive(cfg['definitive'], locations_as_identity(L, ident_ids))
    used = set(pd.to_numeric(s.DHC_Id, errors='coerce').dropna().astype('int64'))
    un = d[~d.DHC_Id.isin(used)][['DHC_Id', 'DHC_Entity_Type', 'DHC_Name',
                                  'DHC_Addr1', 'DHC_City', 'DHC_State',
                                  'DHC_Zip']]
    sheet_data(wb, 'Unreferenced_Definitive',
               un.sort_values(['DHC_Entity_Type', 'DHC_State', 'DHC_Name']))

    wb.save(a.out)
    print(f'Wrote {a.out}')
    for ws in wb.worksheets:
        print(f'  {ws.title:26} {ws.max_row - 1:>7,} rows')


if __name__ == '__main__':
    main()
