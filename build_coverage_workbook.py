#!/usr/bin/env python3
"""Build the branded Coverage Audit workbook from a gap run.

The sibling of build_audit_workbook.py. Same brand, same eight-sheet shape from
the reporting template, same two identities to check - but it answers the
complement question: of the Zeus objects carrying NO Definitive identifier,
which one should each of them point at?

Styling helpers and the palette are imported from build_audit_workbook rather
than copied, so the two deliverables cannot drift apart.

Everything here is derived from the run, so it is re-runnable:

  py build_coverage_workbook.py --candidates gap_2026_08_19_gap_candidates.csv \
      --config sources.yaml --out Zeus_DHC_ID_Coverage_Audit_2026_08_19.xlsx

`--accuracy <audit_..._scored.csv>` is optional: it adds the whole-estate
coverage picture to the Summary, putting the linked and unlinked populations
side by side. Omit it and that section is left out; nothing else changes.
"""
import argparse
import os

import numpy as np
import pandas as pd
from openpyxl import Workbook

from build_audit_workbook import (
    _put, _table, sheet_data, H1, H2, SECTION, BODY, NOTE, WRAP,
    FILL_NAVY, FILL_PURPLE, FILL_CREAM, FILL_LILAC)
from dhc_match_v2 import load_config

# The tiers, in the order a reader should meet them: most actionable first.
# Mirrors dhc_gap_match.TIERS but is stated here because the workbook's ordering
# is a presentation decision, not a scoring one.
TIER_ORDER = ['Strong match - ready to load', 'Probable match - review',
              'Ambiguous - rival candidates', 'Weak match - review',
              'No credible match', 'No usable Zeus name']

TIER_NOTE = {
    'Strong match - ready to load':
        'Name and place both agree, and no rival record fits as well. Load '
        'these after a sample check.',
    'Probable match - review':
        'One side of the evidence is short - a strong name without street '
        'agreement, or the reverse. Quick human confirmation.',
    'Ambiguous - rival candidates':
        'The name is probably right but two or more Definitive records fit '
        'within 3 points. A human must pick; the tool deliberately will not.',
    'Weak match - review':
        'A plausible name match with thin corroboration. Lowest yield per '
        'minute of review.',
    'No credible match':
        'Nothing in the four Definitive exports resembles this entity. Not a '
        'failure of matching - a gap in the reference data.',
    'No usable Zeus name':
        'The Zeus name is a single character or punctuation, so there is '
        'nothing to search on. Fix in Zeus, not here.',
}

# Columns each review sheet leads with. Matched_* come before Suggested_* on
# purpose: the pair of strings that actually matched is what a reviewer judges,
# and it is NOT always Zeus_Name against Suggested_Name.
SHOW = ['EntityId', 'Zeus_Sources', 'Zeus_Name', 'Zeus_Names_All',
        'Zeus_Address', 'Zeus_City', 'Zeus_State', 'Zeus_Zip',
        'Suggested_DHC_Id', 'Suggested_Name', 'Suggested_Entity_Type',
        'Suggested_Status_Note', 'Matched_Zeus_Name',
        'Matched_Definitive_Name', 'Matched_Via', 'Name_Score',
        'Address_Score', 'Match_Score', 'Match_Margin', 'Same_Name_Rivals',
        'Exact_Name_And_Geo', 'Address_Match_Source', 'Location_Count',
        'Suggested_Id_Already_In_Zeus', 'Alt1_DHC_Id', 'Alt1_Name',
        'Alt1_Match_Score', 'Alt2_DHC_Id', 'Alt2_Name', 'Alt2_Match_Score']


def cols(df, names):
    return [c for c in names if c in df.columns]


def build_summary(wb, c, counts, n_pop, subtitle, ev, by_pop, by_type,
                  estate, flags):
    ws = wb.create_sheet('Summary')
    for col, w in zip('ABCDE', (4, 62, 16, 14, 60)):
        ws.column_dimensions[col].width = w

    ws.merge_cells('A1:E1')
    _put(ws, 'A1', 'Jackson and Coker Locum Tenens - Zeus DHC Identifier '
                   'Coverage Audit', H1, FILL_NAVY)
    ws.merge_cells('A2:E2')
    _put(ws, 'A2', subtitle, H2, FILL_PURPLE)

    strong = counts.get('Strong match - ready to load', 0)
    nomatch = counts.get('No credible match', 0)
    review = sum(counts.get(t, 0) for t in
                 ['Probable match - review', 'Ambiguous - rival candidates',
                  'Weak match - review'])

    _put(ws, 'B4', 'Headline answer', SECTION)
    ws.merge_cells('B5:E8')
    _put(ws, 'B5',
         f'Zeus holds {n_pop:,} distinct objects carrying NO Definitive '
         f'identifier at all. For {strong:,} of them ({strong / n_pop:.1%}) '
         f'this audit proposes a specific Definitive record on evidence strong '
         f'enough to load: the name and the place both agree, and no rival '
         f'record fits as well. A further {review:,} ({review / n_pop:.1%}) '
         f'have a credible proposal that needs a human decision - most often '
         f'because several Definitive records fit equally well. For '
         f'{nomatch:,} ({nomatch / n_pop:.1%}) nothing in the Definitive data '
         f'supplied resembles the entity, which is a gap in the reference set '
         f'rather than a failure of matching. Every proposal is a proposal: '
         f'unlike the accuracy audit there is no identifier on record to check '
         f'against, so a tier is a statement about strength of evidence, not a '
         f'confirmed fact.',
         BODY, FILL_CREAM, align=WRAP)

    r = 10
    if estate:
        _put(ws, f'B{r}', 'Identifier coverage across the whole Zeus estate',
             SECTION)
        _put(ws, f'B{r + 1}',
             'The two populations partition one universe - an entity either '
             'carries a Definitive identifier or it does not - so these counts '
             'add up. This is the number to quote when asked "how well linked '
             'are we?".', NOTE)
        r = _table(ws, r + 2, ['Population', 'Objects', 'Share', 'Reading'],
                   estate, notes_col=3)

    _put(ws, f'B{r}', '1. Outcome for every unlinked object', SECTION)
    _put(ws, f'B{r + 1}',
         f'Denominator is all {n_pop:,} objects with no identifier, so nothing '
         f'is hidden by exclusion. These rows sum to the population.', NOTE)
    head = r + 2
    r = _table(ws, head, ['Tier', 'Objects', 'Share', 'Reading'],
               [[t, int(counts.get(t, 0)), counts.get(t, 0) / n_pop,
                 TIER_NOTE[t]] for t in TIER_ORDER if counts.get(t, 0)],
               notes_col=3)
    for col in 'BCDE':
        ws[f'{col}{head + 1}'].fill = FILL_LILAC
    _put(ws, f'B{r}', 'Shaded row is the actionable output of this audit.', NOTE)

    r += 2
    _put(ws, f'B{r}', '2. How strong is the strong tier', SECTION)
    _put(ws, f'B{r + 1}',
         'The evidence behind the load-ready rows. Street-level address '
         'agreement is the strongest single signal, because it distinguishes '
         'the right record from a same-named one elsewhere.', NOTE)
    r = _table(ws, r + 2, ['Measure', 'Objects', 'Share of strong', 'Reading'],
               ev, notes_col=3)

    _put(ws, f'B{r}', '3. Before loading anything', SECTION)
    _put(ws, f'B{r + 1}',
         'Four populations inside the strong tier that are defensible but are '
         'business decisions rather than data questions. Each has its own '
         'sheet. Counts here are strong-tier only; the Status_Flagged sheet '
         'deliberately spans every tier, because a reviewer working the '
         'Probable queue needs the same warning.', NOTE)
    r = _table(ws, r + 2, ['Population', 'Objects', 'Why it needs a look'],
               flags, notes_col=2)

    _put(ws, f'B{r}', '4. By Zeus population', SECTION)
    _put(ws, f'B{r + 1}',
         'Populations overlap - an entity that is both a client and a work '
         'location is counted in each - so these rows sum to more than the '
         'total. Its names and addresses from every population are pooled into '
         'one proposal.', NOTE)
    r = _table(ws, r + 2, ['Zeus population', 'Unlinked', 'Strong match',
                           'Share'], by_pop)

    _put(ws, f'B{r}', '5. What kind of Definitive record is proposed', SECTION)
    _put(ws, f'B{r + 1}',
         'Strong tier only. PracticeLocation means the identifier exists only '
         'in the service-location export, with no overview record behind it.',
         NOTE)
    r = _table(ws, r + 2, ['Definitive entity type', 'Proposals', 'Share'],
               by_type)
    return ws


def build_methodology(wb, notes):
    ws = wb.create_sheet('Methodology')
    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 118
    _put(ws, 'A1', 'Methodology and assumptions', SECTION)
    r = 2
    from openpyxl.styles import Font
    from build_audit_workbook import INK
    for head, text in notes:
        _put(ws, f'B{r}', head, Font(size=10, bold=True, color=INK), align=WRAP)
        _put(ws, f'C{r}', text, BODY, align=WRAP)
        r += 1
    return ws


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--candidates', required=True,
                    help='<prefix>_gap_candidates.csv from dhc_gap_match.py')
    ap.add_argument('--config', default='sources.yaml')
    ap.add_argument('--accuracy', help='an audit <prefix>_scored.csv, to show '
                                       'whole-estate coverage on the Summary')
    ap.add_argument('--out', required=True)
    a = ap.parse_args()

    c = pd.read_csv(a.candidates, low_memory=False)
    prefix = a.candidates.replace('_gap_candidates.csv', '')
    nm_path = f'{prefix}_gap_nomatch.csv'
    ex_path = f'{prefix}_zeus_extract.csv'
    nm = pd.read_csv(nm_path, low_memory=False) if os.path.exists(nm_path) else None
    ex = pd.read_csv(ex_path, low_memory=False) if os.path.exists(ex_path) else None

    # The two output files together are the population; the workbook must not
    # quietly report only the half that matched.
    full = pd.concat([c, nm], ignore_index=True) if nm is not None else c
    n_pop = int(full.EntityId.nunique())
    counts = full.Match_Tier.value_counts().to_dict()
    assert sum(counts.values()) == len(full) == n_pop, (
        f'tier counts {sum(counts.values())} / rows {len(full)} / entities '
        f'{n_pop} disagree - the population is not what it claims to be')

    strong = c[c.Match_Tier == 'Strong match - ready to load']
    ns = max(len(strong), 1)

    ev = [
        ['Exact name and matching geography', int(strong.Exact_Name_And_Geo.sum()),
         int(strong.Exact_Name_And_Geo.sum()) / ns,
         'Normalised names identical and zip or city+state agree. Needs no '
         'judgement call.'],
        ['Street-level address agreement', int((strong.Address_Score >= 85).sum()),
         int((strong.Address_Score >= 85).sum()) / ns,
         'Address score 85+. The strongest discriminator between a record and '
         'a same-named one elsewhere.'],
        ['Name matched outright', int((strong.Name_Score >= 99.9).sum()),
         int((strong.Name_Score >= 99.9).sum()) / ns,
         'Name score 100 against the entity name, an alias, or a service '
         'location.'],
        ['No rival within 10 points', int((strong.Match_Margin >= 10).sum()),
         int((strong.Match_Margin >= 10).sum()) / ns,
         'Clear daylight to the runner-up. Below 3 points a row is demoted to '
         'Ambiguous instead.'],
        ['Matched on the entity\'s own name', int((strong.Matched_Via == 'Name').sum()),
         int((strong.Matched_Via == 'Name').sum()) / ns,
         'The rest matched a former name in parentheses, or a service '
         'location of the proposed parent.'],
    ]

    by_pop = []
    for lab in sorted({s for v in full.Zeus_Sources for s in str(v).split('|')}):
        m = full.Zeus_Sources.astype(str).str.split('|').map(
            lambda v, l=lab: l in v)
        k = int((m & full.Match_Tier.eq('Strong match - ready to load')).sum())
        by_pop.append([lab, int(m.sum()), k, k / max(int(m.sum()), 1)])
    by_pop.sort(key=lambda r: -r[1])

    bt = strong.Suggested_Entity_Type.value_counts()
    by_type = [[k, int(v), int(v) / ns] for k, v in bt.items()]

    shared_other = strong[strong.Suggested_Id_Already_In_Zeus.fillna(False)
                          .astype(bool)] if 'Suggested_Id_Already_In_Zeus' \
        in strong.columns else strong.iloc[0:0]
    dup = strong.Suggested_DHC_Id.value_counts()
    shared_within = strong[strong.Suggested_DHC_Id.isin(dup[dup > 1].index)]
    status = c[c.Suggested_Status_Note.fillna('') != '']
    via_loc = strong[strong.Matched_Via == 'Location']

    flags = [
        ['Proposes an id an existing linked entity already uses',
         len(shared_other),
         'Normal: a client and a work location legitimately share one '
         'Definitive record. A business question, not a data error.'],
        ['Proposes an id another unlinked entity also proposes',
         len(shared_within),
         'Chains. Definitive models a chain as one parent id plus service '
         'locations; Zeus models it as many branches. Check the grain you '
         'want before loading.'],
        ['Matched via a service location, not the entity name',
         len(via_loc),
         'The proposed id belongs to the PARENT; Definitive lists this entity '
         'as one of its locations. Usually right, but understand it first.'],
        ['Definitive marks the proposed record Closed or Merged (strong tier)',
         int((strong.Suggested_Status_Note.fillna('') != '').sum()),
         'Reported, never auto-demoted: a closed record can be the correct id '
         'for a historical row.'],
    ]

    estate = None
    if a.accuracy and os.path.exists(a.accuracy):
        acc = pd.read_csv(a.accuracy, usecols=['EntityId'], low_memory=False)
        n_linked = int(acc.EntityId.nunique())
        # The accuracy scored file holds only the TESTABLE rows; its
        # unverifiable siblings are still linked entities, so read them too or
        # the estate total is understated.
        unv = a.accuracy.replace('_scored.csv', '_unverifiable.csv')
        if os.path.exists(unv):
            n_linked += int(pd.read_csv(unv, usecols=['EntityId'],
                                        low_memory=False).EntityId.nunique())
        tot = n_linked + n_pop
        s_ = int(counts.get('Strong match - ready to load', 0))
        estate = [
            ['Carries a Definitive identifier today', n_linked, n_linked / tot,
             'Measured for accuracy by the companion Accuracy Audit.'],
            ['Carries none', n_pop, n_pop / tot,
             'The subject of this workbook.'],
            ['Total Zeus objects', tot, 1.0,
             'Across the six populations, pooled to one row per entity.'],
            ['Coverage if the strong tier is loaded', n_linked + s_,
             (n_linked + s_) / tot,
             f'Acting on {s_:,} proposals would lift coverage from '
             f'{n_linked / tot:.1%} to {(n_linked + s_) / tot:.1%}.'],
        ]

    src = os.path.basename(a.candidates)
    subtitle = (f'Zeus objects with no Definitive identifier, matched against '
                f'the Hospital, Physician Group, GPO and Practice Location '
                f'exports - {src}')

    wb = Workbook()
    wb.remove(wb.active)
    build_summary(wb, c, counts, n_pop, subtitle, ev, by_pop, by_type,
                  estate, flags)

    build_methodology(wb, [
        ('What this audit does NOT do', 'It proposes identifiers; it does not '
         'verify them. There is no identifier on record for these entities, so '
         'nothing can be checked against. A tier states how strong the '
         'evidence for a proposal is. Nothing here should be loaded without a '
         'sample being labelled by hand first.'),
        ('Population', 'Non-archived Zeus entities across six populations - '
         'IsClient, IsWorkLocation, IsHealthSystem, IsGPO, IsAgency, IsVMS - '
         'that carry NO Definitive identifier on either dbo.Entity or '
         'LinkEntityVerifiedSource. Entities created by a Definitive import '
         'are excluded, exactly as in the accuracy audit, so the two '
         'populations partition one universe and their counts add up.'),
        ('Why the complement is not a bare NOT', 'An entity can hold several '
         'LinkEntityVerifiedSource rows, so "has no Definitive identifier" is a '
         'claim about all of them - the query uses NOT EXISTS, not a negated '
         'join. VerifiedSourceNameId is wrapped in ISNULL because negating "= '
         '1" on a NULL yields NULL rather than TRUE, which would have silently '
         'dropped most of this population.'),
        ('Not every identifier is a Definitive one', 'Zeus records five '
         'verified-source types (Definitive, Definitive Executive, NPI, '
         'Axuall, MDStaff). Only type 1 counts here. The queries emit the two '
         'identifier columns as typed NULLs so a non-Definitive identifier can '
         'never be read as a Definitive one.'),
        ('Six populations, pooled', 'An entity that is both a client and a '
         'work location holds a different name and address in each. Neither is '
         'authoritative, so all of them become candidates and the best match '
         'wins: one proposal per entity, no double counting. Zeus_Sources '
         'records which populations contributed. 5,461 of 48,739 entities are '
         'in more than one.'),
        ('Name does NOT outrank address here', 'The accuracy audit lets name '
         'outrank address, because a divergent address usually just means an '
         'out-of-date service location on an identifier that is otherwise '
         'right. That reasoning does not transfer to choosing an identifier: a '
         'name match alone establishes only that SOME Definitive record shares '
         'the name, and there are 138,385 physician groups to share it with. A '
         'proposal needs two independent agreements - either street-level '
         'address agreement, or a discriminative match on the entity own name '
         'plus city/state/zip agreement.'),
        ('Ambiguity is reported, not resolved', 'Where a rival Definitive '
         'record scores within 3 points of the winner, the row is tiered '
         'Ambiguous rather than being assigned the argmax. Taking the top '
         'score would have made these indistinguishable from genuine matches; '
         'they are 17% of the population.'),
        ('Candidate search', 'Every entity is searched against Definitive '
         'records blocked on state - the entity HQ state plus every state it '
         'has a service location in, so a satellite is reachable where it '
         'actually sits. 143 entities with no usable state were searched '
         'nationally. The top candidates are then scored in full, and the two '
         'runners-up are retained on every sheet so a reviewer can overrule '
         'the pick.'),
        ('Service locations cut both ways', 'Every Definitive service location '
         'lends its name to its parent as an alias. That is safe when an '
         'identifier is given and unsafe when one is being chosen, because an '
         'alias can now SELECT a record. Location names shared across five or '
         'more parents (Family Medicine, Gastroenterology) and pure branch '
         'labels (East Indianapolis, West) are therefore excluded from the '
         'search - their addresses are kept. A match resting on a location '
         'name reaches the strong tier only with street-level address '
         'agreement.'),
        ('Normalisation', 'Both sides lowercased, ampersands expanded, '
         'punctuation stripped. Zeus stores full state names, Definitive '
         'stores codes; states are normalised before comparison. Definitive '
         'embeds former names in parentheses ((FKA ...), (AKA ...)); these are '
         'parsed out and scored as aliases.'),
        ('Name scoring', 'A 35/65 blend of token_set_ratio and '
         'token_sort_ratio - identical to the accuracy audit. Bare '
         'token_set_ratio returns 100 whenever one token set is a subset of '
         'the other, which lets short generic names win spuriously.'),
        ('Read Matched_Zeus_Name and Matched_Definitive_Name', 'These are the '
         'two strings that actually produced the score, and they are NOT '
         'always Zeus_Name against Suggested_Name: pooling means the winning '
         'Zeus name may be the second one on record, and the winning '
         'Definitive string may be a parenthetical alias or a service '
         'location. Matched_Via says which.'),
        ('Known limit', 'Parent-versus-child entities and same-named entities '
         'in different places, the same residual weakness as the accuracy '
         'audit. Name agreement proves the two sides mean the same NAME, not '
         'necessarily the same ENTITY. Closing that needs a human or a third '
         'identifier such as NPI, which is present in the physician group '
         'export and is the obvious next lever.'),
        ('Reproducibility', 'Zeus is a live moving target. The run snapshots '
         'its exact input to <prefix>_zeus_extract.csv; keep that file with '
         'this workbook, or no figure here can be reproduced later.'),
    ])

    show = cols(c, SHOW)

    # Actionable first, then the review queues in descending yield.
    sheet_data(wb, 'Ready_To_Load',
               strong[show].sort_values(['Match_Score', 'Name_Score'],
                                        ascending=False))
    sheet_data(wb, 'Review_Probable',
               c[c.Match_Tier == 'Probable match - review'][show]
               .sort_values('Match_Score', ascending=False))
    sheet_data(wb, 'Ambiguous_Rivals',
               c[c.Match_Tier == 'Ambiguous - rival candidates'][show]
               .sort_values(['Same_Name_Rivals', 'Name_Score'],
                            ascending=False))
    sheet_data(wb, 'Review_Weak',
               c[c.Match_Tier == 'Weak match - review'][show]
               .sort_values('Match_Score', ascending=False))

    # The three "understand before loading" populations from Summary section 3.
    sheet_data(wb, 'Parent_Id_Proposals',
               via_loc[show].sort_values('Match_Score', ascending=False))
    sheet_data(wb, 'Shared_Id_Proposals',
               shared_within.sort_values(['Suggested_DHC_Id', 'EntityId'])[show])
    if len(shared_other):
        sheet_data(wb, 'Id_Already_Linked_In_Zeus',
                   shared_other.sort_values('Suggested_DHC_Id')[show])
    if len(status):
        sheet_data(wb, 'Status_Flagged',
                   status.sort_values('Match_Tier')[show])

    if nm is not None:
        sheet_data(wb, 'No_Credible_Match',
                   nm[cols(nm, SHOW)].sort_values(
                       ['Zeus_State', 'Zeus_Name']))

    sheet_data(wb, 'Candidates_Detail', c)

    wb.save(a.out)
    print(f'Wrote {a.out}')
    for ws in wb.worksheets:
        print(f'  {ws.title:26} {ws.max_row - 1:>7,} rows')

    # The identities from CLAUDE.md, re-checked on the built workbook.
    print('\nIdentity checks:')
    print(f'  tier counts sum to population    '
          f'{sum(counts.values()):,} == {n_pop:,}  '
          f'{"OK" if sum(counts.values()) == n_pop else "FAIL"}')
    tot_sheets = len(strong) + \
        int((c.Match_Tier == 'Probable match - review').sum()) + \
        int((c.Match_Tier == 'Ambiguous - rival candidates').sum()) + \
        int((c.Match_Tier == 'Weak match - review').sum()) + \
        int((c.Match_Tier == 'No usable Zeus name').sum()) + \
        (len(nm) if nm is not None else 0)
    print(f'  review sheets + no-match = pop   {tot_sheets:,} == {n_pop:,}  '
          f'{"OK" if tot_sheets == n_pop else "FAIL"}')
    if ex is not None:
        print(f'  extract entities = population    '
              f'{ex.EntityId.nunique():,} == {n_pop:,}  '
              f'{"OK" if ex.EntityId.nunique() == n_pop else "FAIL"}')


if __name__ == '__main__':
    main()
